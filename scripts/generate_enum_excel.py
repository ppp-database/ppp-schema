"""
enums/*.schema.json から docs/exports/enum.xlsx の各シートを再生成する。

更新対象外（手動管理）:
  - 説明 シート
  - 土地の用途 (LandUsage) シート
  - 不具合現象 (Phenomenon) シート

実行前に docs/exports/backups/{YYYYMMDD}_enum.xlsx へバックアップを作成する。
"""
import json
import shutil
from datetime import date
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
ENUMS_DIR = REPO_ROOT / "enums"
EXPORTS_DIR = REPO_ROOT / "docs" / "exports"
BACKUPS_DIR = EXPORTS_DIR / "backups"
EXCEL_PATH = EXPORTS_DIR / "enum.xlsx"

# enum ファイル名 → Excel シート名 (更新対象のみ)
SHEET_MAP: dict[str, str] = {
    "OrganizationCategory":       "組織種別 (OrganizationCategory)",
    "BuildingCategory":           "建物の種別 (BuildingCategory)",
    "BuildingUsage":              "建物の用途 (BuildingUsage)",
    "LandUseZone":                "土地の用途地域 (LandUseZone)",
    "ContactPointType":           "連絡先の種類 (ContactPointType)",
    "ControlledProperty":         "制御対象 (ControlledProperty)",
    "IdentificationType":         "ID種別 (IdentificationType)",
    "BuildingComponent":          "施設の部位 (BuildingComponent)",
    "Cause":                      "不具合原因 (Cause)",
    "DamageControl":              "簡易処置 (DamageControl)",
    "ReportCategory":             "報告書種別 (ReportCategory)",
    "ComplaintSeverity":          "不具合の緊急性 (ComplaintSeverity)",
    "ComplaintStatus":            "案件のステータス (ComplaintStatus)",
    "ComplaintStep":              "案件のステップ (ComplaintStep)",
}

# BuildingComponent の $def 名 → 表示カテゴリ名 (処理順序を保持)
BC_CATEGORIES: dict[str, str] = {
    "BuildingPartsEnum":               "建物の部位",
    "OutdoorPartsEnum":                "建物外の部位",
    "ElectricalEquipmentEnum":         "電気設備",
    "MechanicalEquipmentEnum":         "機械設備",
    "DisasterPreventionEquipmentEnum": "防災設備",
    "ConveyingEquipmentEnum":          "搬送設備",
    "KitchenEquipmentEnum":            "厨房設備",
    "GasEquipmentEnum":                "ガス設備",
    "ObservationTargetEnum":           "観測対象",
    "OtherEquipmentEnum":              "その他設備",
}

BC_BUILDING_DEFS = {"BuildingPartsEnum", "OutdoorPartsEnum"}


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def flatten_hierarchy(node: dict, path: tuple = ()) -> list[tuple]:
    """x-enumDescriptions を再帰的にフラット化。
    戻り値: [(path_tuple, leaf_value), ...]
    """
    results = []
    for key, value in node.items():
        current = path + (key,)
        if isinstance(value, dict):
            results.extend(flatten_hierarchy(value, current))
        else:
            results.append((current, value))
    return results


def clear_sheet(ws) -> None:
    """シートの全行を削除する。結合セルを先に解除してから行を削除する。"""
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row + 1)


# ---------------------------------------------------------------------------
# シンプルな enum シート (1〜2列)
# ---------------------------------------------------------------------------

def write_simple_sheet(ws, def_schema: dict) -> None:
    """シンプルな enum の1〜2列シートを再生成する。

    x-enumDescriptions 値が文字列 (flat) の場合は 2列: 用語 | 定義
    それ以外は 1列: 用語のみ
    open enum (enum 配列なし) の場合は x-enumDescriptions のキーを用語として使う
    ヘッダ行は出力しない (既存フォーマットに合わせる)
    """
    clear_sheet(ws)

    enum_values: list[str] = def_schema.get("enum", [])
    descriptions: dict = def_schema.get("x-enumDescriptions", {})

    sample = next(iter(descriptions.values()), None) if descriptions else None
    is_flat = isinstance(sample, str)

    if is_flat:
        # 2列: x-enumDescriptions のキー順で出力し、enum にあって descriptions にないものを後置
        written: set[str] = set()
        row_idx = 1
        for term, desc in descriptions.items():
            ws.cell(row_idx, 1, term)
            ws.cell(row_idx, 2, desc)
            written.add(term)
            row_idx += 1
        for term in enum_values:
            if term not in written:
                ws.cell(row_idx, 1, term)
                row_idx += 1
    else:
        # 1列: enum 値。open enum の場合は x-enumDescriptions のキーを使う
        terms = enum_values if enum_values else list(descriptions.keys())
        for row_idx, term in enumerate(terms, 1):
            ws.cell(row_idx, 1, term)


# ---------------------------------------------------------------------------
# BuildingComponent 5パネル シート
# ---------------------------------------------------------------------------

def get_panel_a(defs: dict) -> list[str]:
    """Panel A (col 1): カテゴリ名一覧"""
    return [cat for def_name, cat in BC_CATEGORIES.items() if def_name in defs]


def get_panel_b(defs: dict) -> list[tuple[str, str]]:
    """Panel B (cols 3-4): (カテゴリ, サブカテゴリ) ペア一覧"""
    rows = []
    for def_name, cat_name in BC_CATEGORIES.items():
        if def_name not in defs:
            continue
        desc = defs[def_name].get("x-enumDescriptions", {})
        for level1 in desc:
            rows.append((cat_name, level1))
    return rows


def get_panel_c(defs: dict) -> list[tuple[str, str]]:
    """Panel C (cols 6-7): 用語(建物) — BuildingPartsEnum + OutdoorPartsEnum のリーフ
    戻り値: [(sub_cat, term), ...]  level2="" の3階層のみ
    """
    rows = []
    for def_name in ("BuildingPartsEnum", "OutdoorPartsEnum"):
        if def_name not in defs:
            continue
        desc = defs[def_name].get("x-enumDescriptions", {})
        for path, _ in flatten_hierarchy(desc):
            if len(path) == 3 and path[1] == "":
                rows.append((path[0], path[2]))
    return rows


def get_panel_d_e(defs: dict) -> tuple[list[tuple], list[tuple]]:
    """Panel D (cols 9-10) / E (cols 12-13): 用語(設備)

    Panel D: サブカテゴリ + 分類用語(level2!="") or 直接用語(level2="")
    Panel E: 分類用語 + 用語 (level2!="" のリーフのみ)
    """
    panel_d: list[tuple[str, str]] = []
    panel_e: list[tuple[str, str]] = []
    seen_groups: set[tuple] = set()

    for def_name in BC_CATEGORIES:
        if def_name in BC_BUILDING_DEFS or def_name not in defs:
            continue
        desc = defs[def_name].get("x-enumDescriptions", {})
        for path, _ in flatten_hierarchy(desc):
            if len(path) != 3:
                continue
            level1, level2, level3 = path
            if level2 == "":
                panel_d.append((level1, level3))
            else:
                group_key = (def_name, level1, level2)
                if group_key not in seen_groups:
                    seen_groups.add(group_key)
                    panel_d.append((level1, level2))
                panel_e.append((level2, level3))

    return panel_d, panel_e


def write_bc_sheet(ws, schema: dict) -> None:
    """施設の部位 (BuildingComponent) の5パネルシートを再生成する。

    列レイアウト:
      col 1  : Panel A  カテゴリ
      col 2  : spacer
      col 3-4: Panel B  サブカテゴリ (カテゴリ | サブカテゴリ)
      col 5  : spacer
      col 6-7: Panel C  用語(建物)  (サブカテゴリ | 用語)
      col 8  : spacer
      col 9-10: Panel D 用語(設備)flat   (サブカテゴリ | 用語(分類用語))
      col 11 : spacer
      col 12-13: Panel E 用語(設備)hier  (分類用語 | 用語)
    """
    clear_sheet(ws)

    defs = schema.get("$defs", {})

    panel_a = get_panel_a(defs)
    panel_b = get_panel_b(defs)
    panel_c = get_panel_c(defs)
    panel_d, panel_e = get_panel_d_e(defs)

    # 行1: パネル見出し
    ws.cell(1, 1, "カテゴリ")
    ws.cell(1, 3, "サブカテゴリ")
    ws.cell(1, 6, "用語(建物)")
    ws.cell(1, 9, "用語(設備)")
    ws.cell(1, 12, "用語(設備)")

    # 行2: 列見出し
    ws.cell(2, 3, "カテゴリ")
    ws.cell(2, 4, "サブカテゴリ")
    ws.cell(2, 6, "サブカテゴリ")
    ws.cell(2, 7, "用語(分類用語)")
    ws.cell(2, 9, "サブカテゴリ")
    ws.cell(2, 10, "用語(分類用語)")
    ws.cell(2, 12, "分類用語")
    ws.cell(2, 13, "用語")

    # データ行 (3行目から)
    n = max(len(panel_a), len(panel_b), len(panel_c), len(panel_d), len(panel_e), 0)
    for i in range(n):
        row = i + 3
        if i < len(panel_a):
            ws.cell(row, 1, panel_a[i])
        if i < len(panel_b):
            ws.cell(row, 3, panel_b[i][0])
            ws.cell(row, 4, panel_b[i][1])
        if i < len(panel_c):
            ws.cell(row, 6, panel_c[i][0])
            ws.cell(row, 7, panel_c[i][1])
        if i < len(panel_d):
            ws.cell(row, 9, panel_d[i][0])
            ws.cell(row, 10, panel_d[i][1])
        if i < len(panel_e):
            ws.cell(row, 12, panel_e[i][0])
            ws.cell(row, 13, panel_e[i][1])


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def get_primary_def(schema: dict) -> dict | None:
    """$defs から主要 def (oneOf 合成ではない最初のもの) を返す。"""
    for def_schema in schema.get("$defs", {}).values():
        if "oneOf" in def_schema and "enum" not in def_schema and not def_schema.get("x-enumDescriptions"):
            continue
        return def_schema
    return None


def backup() -> Path:
    """既存の Excel を backups/ へコピーする。"""
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    dst = BACKUPS_DIR / f"{today}_enum.xlsx"
    shutil.copy2(EXCEL_PATH, dst)
    print(f"[backup] {dst}")
    return dst


def main() -> None:
    backup()

    wb = openpyxl.load_workbook(EXCEL_PATH)

    for enum_name, sheet_name in SHEET_MAP.items():
        schema_path = ENUMS_DIR / f"{enum_name}.schema.json"
        if not schema_path.exists():
            print(f"[skip] {enum_name}: schema not found")
            continue
        if sheet_name not in wb.sheetnames:
            print(f"[skip] {sheet_name}: sheet not in workbook")
            continue

        schema = json.loads(schema_path.read_text(encoding="utf-8"))
        ws = wb[sheet_name]

        if enum_name == "BuildingComponent":
            write_bc_sheet(ws, schema)
        else:
            def_schema = get_primary_def(schema)
            if def_schema is None:
                print(f"[skip] {enum_name}: no primary def found")
                continue
            write_simple_sheet(ws, def_schema)

        print(f"[ok] {sheet_name}")

    wb.save(EXCEL_PATH)
    print(f"[saved] {EXCEL_PATH}")


if __name__ == "__main__":
    main()
